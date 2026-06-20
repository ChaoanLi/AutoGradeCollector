#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
智能成绩统计脚本

功能：
- 自动发现多个名单文件（平时成绩登记表、CL 文件）
- 自动发现 SA/LA/TL 成绩文件
- 自动识别成绩列与满分，并记录识别来源
- 按名单合并学生，缺失成绩按 0 处理但保留缺失标记
- 输出成绩统计留存表，不覆盖原始登记表
"""

from __future__ import annotations

import argparse
import math
import re
from dataclasses import dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import pandas as pd
import xlrd
from openpyxl.utils import get_column_letter
from xlwt.Cell import NumberCell
from xlutils.copy import copy as copy_xlrd_workbook


COMMON_FULL_SCORES = (5, 10, 20, 30, 50, 100)
DEFAULT_DISCUSSION_TARGET = 10.0
DEFAULT_MOOC_TARGET = 20.0
OUTPUT_FILE = "成绩统计留存表.xlsx"


@dataclass
class Student:
    student_id: str
    name: str = ""
    class_name: str = ""
    gender: str = ""
    course_sequences: set[str] = field(default_factory=set)
    roster_sources: list[str] = field(default_factory=list)
    duplicate_records: list[str] = field(default_factory=list)


@dataclass
class TargetScale:
    discussion: float = DEFAULT_DISCUSSION_TARGET
    discussion_source: str = "默认值"
    mooc: float = DEFAULT_MOOC_TARGET
    mooc_source: str = "默认值"


@dataclass
class ScoreSource:
    file_path: Path
    score_type: str
    label: str
    score_column: str
    full_score: float | None
    full_score_source: str
    status: str
    max_observed: float | None
    records: dict[str, float]
    missing_column_reason: str = ""


def read_excel(path: Path, header: int | None = 0) -> pd.DataFrame:
    suffix = path.suffix.lower()
    try:
        if suffix == ".xls":
            return pd.read_excel(path, header=header, engine="xlrd")
        return pd.read_excel(path, header=header, engine="openpyxl")
    except ImportError as exc:
        if suffix == ".xls":
            raise RuntimeError(
                "读取 .xls 需要安装 xlrd：python -m pip install xlrd"
            ) from exc
        raise


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def is_student_id(value: Any) -> bool:
    text = clean_cell(value)
    if not text:
        return False
    if re.fullmatch(r"[US]\d{6,}", text, flags=re.I):
        return True
    if len(text) >= 8 and any(ch.isdigit() for ch in text):
        return not text.replace(".", "").replace("-", "").isdigit()
    return False


def to_number(value: Any) -> float | None:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        text = clean_cell(value)
        if not text:
            return None
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if not match:
            return None
        number = float(match.group(0))
    if math.isnan(number):
        return None
    return number


def discover_files(data_dir: Path) -> tuple[list[Path], list[Path], list[Path], list[Path]]:
    excel_files = sorted(
        p
        for pattern in ("*.xls", "*.xlsx")
        for p in data_dir.glob(pattern)
        if not p.name.startswith("~$")
    )
    output_names = {OUTPUT_FILE}
    excel_files = [p for p in excel_files if p.name not in output_names]
    excel_files = [p for p in excel_files if "_已填" not in p.stem]

    roster_files = [
        p
        for p in excel_files
        if "平时成绩登记表" in p.name or re.match(r"^CL", p.stem, flags=re.I)
    ]
    sa_files = [p for p in excel_files if re.match(r"^SA[_\-\s]?\d*", p.stem, flags=re.I)]
    la_files = [p for p in excel_files if re.match(r"^LA[_\-\s]?\d*", p.stem, flags=re.I)]
    tl_files = [p for p in excel_files if re.match(r"^TL[_\-\s]?\d*", p.stem, flags=re.I)]
    return roster_files, sa_files, la_files, tl_files


def find_header_mapping(raw: pd.DataFrame) -> tuple[int | None, dict[str, int]]:
    best_row: int | None = None
    best_mapping: dict[str, int] = {}
    for row_idx in range(min(20, len(raw))):
        mapping: dict[str, int] = {}
        for col_idx, value in enumerate(raw.iloc[row_idx].tolist()):
            cell = clean_cell(value).lower()
            if not cell:
                continue
            if "学号" in cell or "student id" in cell:
                mapping["student_id"] = col_idx
            elif "姓名" in cell or "真实姓名" in cell or cell in {"name", "student name"}:
                mapping["name"] = col_idx
            elif "班级" in cell or "class" in cell:
                mapping["class_name"] = col_idx
            elif "性别" in cell or "gender" in cell:
                mapping["gender"] = col_idx
            elif "序号" in cell or cell in {"no", "编号"}:
                mapping["sequence"] = col_idx
        if "student_id" in mapping and "name" in mapping:
            best_row = row_idx
            best_mapping = mapping
            break
    return best_row, best_mapping


def parse_course_sequence(path: Path, raw: pd.DataFrame) -> str:
    name_match = re.search(r"-(\d{3})(?:\D|$)", path.stem)
    if name_match:
        return name_match.group(1)
    for _, row in raw.head(8).iterrows():
        text = " ".join(clean_cell(value) for value in row.tolist())
        match = re.search(r"课序号[:：]?\s*(\d{3})", text)
        if match:
            return match.group(1)
    return ""


def update_target_scale_from_raw(scale: TargetScale, raw: pd.DataFrame, source_name: str) -> None:
    for _, row in raw.head(20).iterrows():
        text = " ".join(clean_cell(value) for value in row.tolist())
        for keyword, attr, source_attr in (
            ("讨论", "discussion", "discussion_source"),
            ("主题讨论", "discussion", "discussion_source"),
            ("慕课", "mooc", "mooc_source"),
        ):
            if keyword not in text:
                continue
            match = re.search(rf"{keyword}[^\d]{{0,8}}(\d+(?:\.\d+)?)\s*分制", text)
            if match:
                setattr(scale, attr, float(match.group(1)))
                setattr(scale, source_attr, f"{source_name} 说明行")


def extract_rosters(roster_files: list[Path]) -> tuple[dict[str, Student], TargetScale, list[dict[str, Any]]]:
    students: dict[str, Student] = {}
    target_scale = TargetScale()
    roster_report: list[dict[str, Any]] = []

    for path in roster_files:
        raw = read_excel(path, header=None)
        update_target_scale_from_raw(target_scale, raw, path.name)
        header_row, mapping = find_header_mapping(raw)
        course_sequence = parse_course_sequence(path, raw)
        extracted = 0

        if header_row is None:
            roster_report.append(
                {
                    "名单文件": path.name,
                    "课序号": course_sequence,
                    "识别状态": "未找到包含学号和姓名的表头",
                    "学生数": 0,
                }
            )
            continue

        for _, row in raw.iloc[header_row + 1 :].iterrows():
            values = row.tolist()
            sid = clean_cell(values[mapping["student_id"]]) if mapping["student_id"] < len(values) else ""
            if not is_student_id(sid):
                continue

            name = clean_cell(values[mapping["name"]]) if mapping["name"] < len(values) else ""
            class_name = ""
            gender = ""
            if "class_name" in mapping and mapping["class_name"] < len(values):
                class_name = clean_cell(values[mapping["class_name"]])
            if "gender" in mapping and mapping["gender"] < len(values):
                gender = clean_cell(values[mapping["gender"]])

            source_record = f"{path.name}"
            if course_sequence:
                source_record += f"(课序号{course_sequence})"

            if sid not in students:
                students[sid] = Student(
                    student_id=sid,
                    name=name,
                    class_name=class_name,
                    gender=gender,
                )
            else:
                students[sid].duplicate_records.append(source_record)
                if not students[sid].name and name:
                    students[sid].name = name
                if not students[sid].class_name and class_name:
                    students[sid].class_name = class_name
                if not students[sid].gender and gender:
                    students[sid].gender = gender

            students[sid].roster_sources.append(source_record)
            if course_sequence:
                students[sid].course_sequences.add(course_sequence)
            extracted += 1

        roster_report.append(
            {
                "名单文件": path.name,
                "课序号": course_sequence,
                "识别状态": "成功",
                "学生数": extracted,
                "表头行": header_row + 1,
            }
        )

    return students, target_scale, roster_report


def score_column_priority(score_type: str, column: Any) -> int | None:
    name = clean_cell(column)
    if not name:
        return None
    excluded = ("第1次", "第2次", "第3次", "互评", "其它互评", "调整")
    if any(token in name for token in excluded):
        return None

    if score_type == "SA":
        if re.search(r"^得分\s*/\s*\d+(?:\.\d+)?\s*分$", name):
            return 0
        if name == "得分":
            return 1
        if "得分" in name and "分" in name:
            return 2
    elif score_type == "LA":
        if name == "得分":
            return 0
        if re.search(r"^得分\s*/\s*\d+(?:\.\d+)?\s*分$", name):
            return 1
    elif score_type == "TL":
        excluded_tl = ("成绩/", "原始分", "测验/", "作业/", "考试/", "视频/", "域外")
        if name in excluded_tl:
            return None
        if name == "讨论/":
            return 0
        if "主题讨论" in name:
            return 1
        if "讨论" in name:
            return 2
    return None


def choose_score_column(df: pd.DataFrame, score_type: str) -> tuple[str | None, str]:
    candidates: list[tuple[int, str]] = []
    for column in df.columns:
        priority = score_column_priority(score_type, column)
        if priority is not None:
            candidates.append((priority, clean_cell(column)))
    if not candidates:
        return None, f"未找到 {score_type} 可用成绩列"
    candidates.sort(key=lambda item: (item[0], item[1]))
    return candidates[0][1], ""


def full_score_from_text(text: str) -> tuple[float | None, str]:
    patterns = (
        r"/\s*(\d+(?:\.\d+)?)\s*分",
        r"(\d+(?:\.\d+)?)\s*分制",
        r"满分\s*(\d+(?:\.\d+)?)",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return float(match.group(1)), "列名"
    return None, ""


def infer_full_score(column_name: str, values: pd.Series, raw: pd.DataFrame, score_type: str) -> tuple[float | None, str, str, float | None]:
    full_score, source = full_score_from_text(column_name)
    numeric_values = pd.to_numeric(values, errors="coerce").dropna()
    max_observed = float(numeric_values.max()) if len(numeric_values) else None
    if full_score is not None:
        return full_score, source, "确认", max_observed

    for _, row in raw.head(20).iterrows():
        text = " ".join(clean_cell(value) for value in row.tolist())
        if score_type == "TL" and "讨论" not in text:
            continue
        if score_type in {"SA", "LA"} and "得分" not in text:
            continue
        full_score, _ = full_score_from_text(text)
        if full_score is not None:
            return full_score, "说明行", "确认", max_observed

    if max_observed is None:
        return None, "无有效数值", "需确认", None

    for score in COMMON_FULL_SCORES:
        if max_observed <= score + 1e-9:
            if max_observed >= score * 0.85 or abs(max_observed - score) < 1e-9:
                return float(score), "分布推断", "确认", max_observed
            return float(score), "分布推断", "需确认", max_observed

    rounded = math.ceil(max_observed / 10.0) * 10.0
    return rounded, "分布推断", "需确认", max_observed


def extract_score_source(path: Path, score_type: str) -> ScoreSource:
    df = read_excel(path, header=0)
    df = df.rename(columns={column: clean_cell(column) for column in df.columns})
    raw = read_excel(path, header=None)
    label = path.stem
    score_column, reason = choose_score_column(df, score_type)
    if score_column is None:
        return ScoreSource(
            file_path=path,
            score_type=score_type,
            label=label,
            score_column="",
            full_score=None,
            full_score_source="",
            status="需确认",
            max_observed=None,
            records={},
            missing_column_reason=reason,
        )

    full_score, full_score_source, status, max_observed = infer_full_score(
        score_column, df[score_column], raw, score_type
    )

    records: dict[str, float] = {}
    student_id_column = find_student_id_column(df)
    if student_id_column is None:
        status = "需确认"
        reason = "未找到学号列"
    else:
        for _, row in df.iterrows():
            sid = clean_cell(row.get(student_id_column))
            if not is_student_id(sid):
                continue
            score = to_number(row.get(score_column))
            if score is None:
                continue
            records[sid] = score

    return ScoreSource(
        file_path=path,
        score_type=score_type,
        label=label,
        score_column=score_column,
        full_score=full_score,
        full_score_source=full_score_source,
        status=status,
        max_observed=max_observed,
        records=records,
        missing_column_reason=reason,
    )


def find_student_id_column(df: pd.DataFrame) -> str | None:
    for column in df.columns:
        name = clean_cell(column).lower()
        if name == "学号" or "student id" in name:
            return column
    for column in df.columns:
        name = clean_cell(column).lower()
        if "学号" in name:
            return column
    return None


def extract_all_scores(sa_files: list[Path], la_files: list[Path], tl_files: list[Path]) -> list[ScoreSource]:
    sources: list[ScoreSource] = []
    for score_type, files in (("SA", sa_files), ("LA", la_files), ("TL", tl_files)):
        for path in files:
            sources.append(extract_score_source(path, score_type))
    return sources


def safe_divide(value: float, denominator: float | None, target: float) -> float:
    if not denominator:
        return 0.0
    return value / denominator * target


def round_half_up_to_int(value: Any) -> int:
    number = to_number(value)
    if number is None:
        return 0
    return int(Decimal(str(number)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def build_detail_rows(
    students: dict[str, Student],
    sources: list[ScoreSource],
    target_scale: TargetScale,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    sa_la_sources = [source for source in sources if source.score_type in {"SA", "LA"}]
    tl_sources = [source for source in sources if source.score_type == "TL"]
    mooc_full = sum(source.full_score or 0.0 for source in sa_la_sources)

    detail_rows: list[dict[str, Any]] = []
    issue_rows: list[dict[str, Any]] = []

    for index, student in enumerate(sorted(students.values(), key=lambda s: (sorted(s.course_sequences), s.student_id)), 1):
        row: dict[str, Any] = {
            "序号": index,
            "课序号": "、".join(sorted(student.course_sequences)),
            "学号": student.student_id,
            "姓名": student.name,
            "班级": student.class_name,
            "性别": student.gender,
            "名单来源": "；".join(dict.fromkeys(student.roster_sources)),
            "重复名单记录": "；".join(student.duplicate_records),
        }
        missing: list[str] = []
        mooc_raw = 0.0

        for source in sa_la_sources:
            score = source.records.get(student.student_id, 0.0)
            if student.student_id not in source.records:
                missing.append(source.label)
            mooc_raw += score
            row[f"{source.label}_原始分"] = round(score, 2)
            row[f"{source.label}_满分"] = source.full_score

        discussion_raw_total = 0.0
        discussion_full_total = 0.0
        for source in tl_sources:
            score = source.records.get(student.student_id, 0.0)
            if student.student_id not in source.records:
                missing.append(source.label)
            discussion_raw_total += score
            discussion_full_total += source.full_score or 0.0
            row[f"{source.label}_讨论原始分"] = round(score, 2)
            row[f"{source.label}_讨论满分"] = source.full_score
            row[f"{source.label}_讨论换算分"] = round(
                safe_divide(score, source.full_score, target_scale.discussion), 4
            )

        row["慕课原始总分"] = round(mooc_raw, 2)
        row["慕课总满分"] = round(mooc_full, 2)
        row["慕课目标分"] = target_scale.mooc
        row["慕课20分制"] = round(safe_divide(mooc_raw, mooc_full, target_scale.mooc), 2)
        row["讨论原始总分"] = round(discussion_raw_total, 2)
        row["讨论总满分"] = round(discussion_full_total, 2)
        row["讨论目标分"] = target_scale.discussion
        row["讨论10分制"] = round(
            safe_divide(discussion_raw_total, discussion_full_total, target_scale.discussion),
            2,
        )
        row["缺失项目数"] = len(missing)
        row["缺失来源清单"] = "；".join(missing)
        detail_rows.append(row)

        if missing or student.duplicate_records:
            issue_rows.append(
                {
                    "学号": student.student_id,
                    "姓名": student.name,
                    "问题类型": "缺失成绩/重复名单",
                    "详情": f"缺失：{'；'.join(missing)}；重复：{'；'.join(student.duplicate_records)}",
                }
            )

    roster_ids = set(students.keys())
    for source in sources:
        for sid, score in source.records.items():
            if sid not in roster_ids:
                issue_rows.append(
                    {
                        "学号": sid,
                        "姓名": "",
                        "问题类型": "成绩源中存在但名单缺失",
                        "详情": f"{source.label}: {score}",
                    }
                )

    return detail_rows, issue_rows


def build_full_score_rows(sources: list[ScoreSource], target_scale: TargetScale) -> list[dict[str, Any]]:
    rows = [
        {
            "文件": source.file_path.name,
            "类型": source.score_type,
            "成绩列": source.score_column,
            "识别满分": source.full_score,
            "满分识别来源": source.full_score_source,
            "观察最高分": source.max_observed,
            "记录数": len(source.records),
            "状态": source.status,
            "备注": source.missing_column_reason,
        }
        for source in sources
    ]
    rows.extend(
        [
            {
                "文件": "目标分制",
                "类型": "讨论",
                "成绩列": "讨论/主题讨论",
                "识别满分": target_scale.discussion,
                "满分识别来源": target_scale.discussion_source,
                "观察最高分": "",
                "记录数": "",
                "状态": "确认" if target_scale.discussion_source != "默认值" else "默认值",
                "备注": "登记表目标列分制",
            },
            {
                "文件": "目标分制",
                "类型": "慕课",
                "成绩列": "SA+LA",
                "识别满分": target_scale.mooc,
                "满分识别来源": target_scale.mooc_source,
                "观察最高分": "",
                "记录数": "",
                "状态": "确认" if target_scale.mooc_source != "默认值" else "默认值",
                "备注": "登记表目标列分制",
            },
        ]
    )
    return rows


def write_output(
    output_path: Path,
    detail_rows: list[dict[str, Any]],
    full_score_rows: list[dict[str, Any]],
    roster_rows: list[dict[str, Any]],
    issue_rows: list[dict[str, Any]],
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(detail_rows).to_excel(writer, sheet_name="成绩明细", index=False)
        pd.DataFrame(full_score_rows).to_excel(writer, sheet_name="满分识别", index=False)
        pd.DataFrame(roster_rows).to_excel(writer, sheet_name="名单来源", index=False)
        pd.DataFrame(issue_rows).to_excel(writer, sheet_name="缺失与重复", index=False)
        apply_detail_formulas(writer.book["成绩明细"])
        try:
            writer.book.calculation.fullCalcOnLoad = True
            writer.book.calculation.forceFullCalc = True
        except AttributeError:
            pass

        for worksheet in writer.book.worksheets:
            worksheet.freeze_panes = "A2"
            for column_cells in worksheet.columns:
                header = clean_cell(column_cells[0].value)
                max_len = max([len(clean_cell(cell.value)) for cell in column_cells[:80]] + [len(header)])
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 32)


def apply_detail_formulas(worksheet: Any) -> None:
    headers = [clean_cell(cell.value) for cell in worksheet[1]]
    col_by_header = {header: idx + 1 for idx, header in enumerate(headers) if header}

    def addr(header: str, row_idx: int) -> str:
        return f"{get_column_letter(col_by_header[header])}{row_idx}"

    def sum_formula(source_headers: list[str], row_idx: int) -> str:
        if not source_headers:
            return "0"
        cells = ",".join(addr(header, row_idx) for header in source_headers)
        return f"SUM({cells})"

    mooc_raw_headers = [
        header
        for header in headers
        if header.endswith("_原始分") and not header.endswith("_讨论原始分")
    ]
    mooc_full_headers = [
        header
        for header in headers
        if header.endswith("_满分") and not header.endswith("_讨论满分")
    ]
    discussion_raw_headers = [header for header in headers if header.endswith("_讨论原始分")]
    discussion_full_headers = [header for header in headers if header.endswith("_讨论满分")]
    discussion_scaled_headers = [header for header in headers if header.endswith("_讨论换算分")]

    for row_idx in range(2, worksheet.max_row + 1):
        if "慕课原始总分" in col_by_header:
            worksheet[addr("慕课原始总分", row_idx)] = f"={sum_formula(mooc_raw_headers, row_idx)}"
        if "慕课总满分" in col_by_header:
            worksheet[addr("慕课总满分", row_idx)] = f"={sum_formula(mooc_full_headers, row_idx)}"
        if {"慕课20分制", "慕课原始总分", "慕课总满分", "慕课目标分"} <= set(col_by_header):
            worksheet[addr("慕课20分制", row_idx)] = (
                f'=IFERROR(ROUND({addr("慕课原始总分", row_idx)}/'
                f'{addr("慕课总满分", row_idx)}*{addr("慕课目标分", row_idx)},2),0)'
            )

        if "讨论原始总分" in col_by_header:
            worksheet[addr("讨论原始总分", row_idx)] = f"={sum_formula(discussion_raw_headers, row_idx)}"
        if "讨论总满分" in col_by_header:
            worksheet[addr("讨论总满分", row_idx)] = f"={sum_formula(discussion_full_headers, row_idx)}"
        if {"讨论10分制", "讨论原始总分", "讨论总满分", "讨论目标分"} <= set(col_by_header):
            worksheet[addr("讨论10分制", row_idx)] = (
                f'=IFERROR(ROUND({addr("讨论原始总分", row_idx)}/'
                f'{addr("讨论总满分", row_idx)}*{addr("讨论目标分", row_idx)},2),0)'
            )

        for header in discussion_scaled_headers:
            prefix = header[: -len("_讨论换算分")]
            raw_header = f"{prefix}_讨论原始分"
            full_header = f"{prefix}_讨论满分"
            if raw_header in col_by_header and full_header in col_by_header and "讨论目标分" in col_by_header:
                worksheet[addr(header, row_idx)] = (
                    f'=IFERROR(ROUND({addr(raw_header, row_idx)}/'
                    f'{addr(full_header, row_idx)}*{addr("讨论目标分", row_idx)},4),0)'
                )

    for header in (
        discussion_scaled_headers
        + ["慕课原始总分", "慕课总满分", "慕课20分制", "讨论原始总分", "讨论总满分", "讨论10分制"]
    ):
        if header in col_by_header:
            for row_idx in range(2, worksheet.max_row + 1):
                worksheet[addr(header, row_idx)].number_format = "0.00"


def find_target_column(raw: pd.DataFrame, keywords: tuple[str, ...]) -> int | None:
    candidates: list[tuple[int, int]] = []
    for row_idx in range(min(12, len(raw))):
        for col_idx, value in enumerate(raw.iloc[row_idx].tolist()):
            text = clean_cell(value)
            if any(keyword in text for keyword in keywords):
                candidates.append((row_idx, col_idx))
    if not candidates:
        return None
    # Prefer header/description rows over inline notes in the first student row.
    candidates.sort(key=lambda item: (item[0] > 5, item[0], item[1]))
    return candidates[0][1]


def fill_cl_workbooks(
    cl_files: list[Path],
    detail_rows: list[dict[str, Any]],
    overwrite: bool = False,
) -> list[Path]:
    scores_by_id = {
        clean_cell(row.get("学号")): {
            "discussion": row.get("讨论10分制", 0),
            "mooc": row.get("慕课20分制", 0),
        }
        for row in detail_rows
    }
    outputs: list[Path] = []

    for path in cl_files:
        raw = read_excel(path, header=None)
        header_row, mapping = find_header_mapping(raw)
        discussion_col = find_target_column(raw, ("主题讨论", "讨论"))
        mooc_col = find_target_column(raw, ("慕课",))

        if header_row is None or "student_id" not in mapping:
            print(f"跳过填表 {path.name}: 未找到学号表头")
            continue
        if discussion_col is None or mooc_col is None:
            print(f"跳过填表 {path.name}: 未找到讨论或慕课列")
            continue

        book = xlrd.open_workbook(str(path), formatting_info=True)
        writable = copy_xlrd_workbook(book)
        sheet = writable.get_sheet(0)
        filled = 0

        def write_number_preserving_style(row_idx: int, col_idx: int, value: int) -> None:
            row = sheet.row(row_idx)
            existing_cell = row._Row__cells.get(col_idx)
            xf_idx = existing_cell.xf_idx if existing_cell is not None else 17
            row.insert_cell(col_idx, NumberCell(row_idx, col_idx, xf_idx, value))

        for row_idx in range(header_row + 1, len(raw)):
            values = raw.iloc[row_idx].tolist()
            sid = clean_cell(values[mapping["student_id"]]) if mapping["student_id"] < len(values) else ""
            if not is_student_id(sid):
                continue
            score = scores_by_id.get(sid, {"discussion": 0, "mooc": 0})
            write_number_preserving_style(
                row_idx,
                discussion_col,
                round_half_up_to_int(score["discussion"]),
            )
            write_number_preserving_style(
                row_idx,
                mooc_col,
                round_half_up_to_int(score["mooc"]),
            )
            filled += 1

        output_path = path if overwrite else path.with_name(f"{path.stem}_已填{path.suffix}")
        save_path = path.with_name(f"{path.stem}.__tmp_filled__{path.suffix}") if overwrite else output_path
        writable.save(str(save_path))
        if overwrite:
            save_path.replace(path)
        outputs.append(output_path)
        print(
            f"{'已覆盖' if overwrite else '已生成'} {output_path.name}: 填入 {filled} 人，"
            f"讨论列第 {discussion_col + 1} 列，慕课列第 {mooc_col + 1} 列"
        )

    return outputs


def run(
    data_dir: Path,
    output_path: Path,
    fill_cl: bool = False,
    overwrite_cl: bool = False,
) -> list[Path]:
    roster_files, sa_files, la_files, tl_files = discover_files(data_dir)
    if not roster_files:
        raise RuntimeError("未找到名单文件：需要 *平时成绩登记表*.xls/xlsx 或 CL*.xls/xlsx")

    print("发现名单文件:")
    for path in roster_files:
        print(f"  - {path.name}")
    print("发现成绩文件:")
    for path in sa_files + la_files + tl_files:
        print(f"  - {path.name}")

    students, target_scale, roster_rows = extract_rosters(roster_files)
    sources = extract_all_scores(sa_files, la_files, tl_files)
    detail_rows, issue_rows = build_detail_rows(students, sources, target_scale)
    full_score_rows = build_full_score_rows(sources, target_scale)
    write_output(output_path, detail_rows, full_score_rows, roster_rows, issue_rows)
    filled_outputs: list[Path] = []
    if fill_cl or overwrite_cl:
        cl_files = [path for path in roster_files if re.match(r"^CL", path.stem, flags=re.I)]
        filled_outputs = fill_cl_workbooks(cl_files, detail_rows, overwrite=overwrite_cl)

    need_confirm = [row for row in full_score_rows if row.get("状态") == "需确认"]
    print(f"\n学生数: {len(students)}")
    print(f"成绩源数: {len(sources)}")
    print(f"输出文件: {output_path}")
    for path in filled_outputs:
        print(f"填表副本: {path}")
    if need_confirm: print("注意：存在需要确认的满分识别项，请查看“满分识别”工作表。")
    return filled_outputs


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="智能成绩统计并生成留存表")
    parser.add_argument("--data-dir", default=".", help="成绩文件所在目录")
    parser.add_argument("--output", default=OUTPUT_FILE, help="输出 xlsx 文件名")
    parser.add_argument("--fill-cl", action="store_true", help="生成填好讨论和慕课列的 CL_*.xls 副本")
    parser.add_argument("--overwrite-cl", action="store_true", help="直接覆盖 CL_*.xls 的讨论和慕课列")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    data_dir = Path(args.data_dir).resolve()
    output_path = Path(args.output)
    if not output_path.is_absolute():
        output_path = data_dir / output_path
    run(data_dir, output_path, fill_cl=args.fill_cl, overwrite_cl=args.overwrite_cl)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
