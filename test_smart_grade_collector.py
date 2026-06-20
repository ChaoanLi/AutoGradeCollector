import tempfile
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from smart_grade_collector import (
    choose_score_column,
    extract_rosters,
    infer_full_score,
    round_half_up_to_int,
    write_output,
)


class SmartGradeCollectorTests(unittest.TestCase):
    def test_sa_prefers_total_score_column_with_full_score(self):
        df = pd.DataFrame(
            {
                "学号": ["U202400001"],
                "得分/100分": [96],
                "第1次得分": [90],
            }
        )
        column, reason = choose_score_column(df, "SA")
        self.assertEqual(reason, "")
        self.assertEqual(column, "得分/100分")

        full_score, source, status, max_observed = infer_full_score(
            column, df[column], df, "SA"
        )
        self.assertEqual(full_score, 100)
        self.assertEqual(source, "列名")
        self.assertEqual(status, "确认")
        self.assertEqual(max_observed, 96)

    def test_distribution_inference_marks_low_max_as_needs_confirmation(self):
        df = pd.DataFrame({"学号": ["U202400001", "U202400002"], "得分": [12, 13]})
        full_score, source, status, max_observed = infer_full_score(
            "得分", df["得分"], df, "LA"
        )
        self.assertEqual(full_score, 20)
        self.assertEqual(source, "分布推断")
        self.assertEqual(status, "需确认")
        self.assertEqual(max_observed, 13)

    def test_tl_prefers_discussion_and_excludes_total_score_columns(self):
        df = pd.DataFrame(
            {
                "学号": ["U202400001"],
                "成绩/": [66],
                "原始分": [66],
                "主题讨论": [8],
            }
        )
        column, reason = choose_score_column(df, "TL")
        self.assertEqual(reason, "")
        self.assertEqual(column, "主题讨论")

    def test_cl_fill_rounding_uses_half_up_integers(self):
        self.assertEqual(round_half_up_to_int(14.49), 14)
        self.assertEqual(round_half_up_to_int(14.5), 15)
        self.assertEqual(round_half_up_to_int("3.5"), 4)

    def test_output_summary_columns_are_formulas(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "out.xlsx"
            write_output(
                output,
                [
                    {
                        "学号": "U202400001",
                        "SA_1_原始分": 50,
                        "SA_1_满分": 50,
                        "LA_1_原始分": 30,
                        "LA_1_满分": 30,
                        "TL_讨论原始分": 30,
                        "TL_讨论满分": 30,
                        "TL_讨论换算分": 0,
                        "慕课原始总分": 0,
                        "慕课总满分": 0,
                        "慕课目标分": 20,
                        "慕课20分制": 0,
                        "讨论原始总分": 0,
                        "讨论总满分": 0,
                        "讨论目标分": 10,
                        "讨论10分制": 0,
                    }
                ],
                [],
                [],
                [],
            )
            ws = load_workbook(output, data_only=False)["成绩明细"]
            headers = [cell.value for cell in ws[1]]
            values = {header: ws.cell(2, idx + 1).value for idx, header in enumerate(headers)}

        self.assertTrue(str(values["慕课原始总分"]).startswith("=SUM("))
        self.assertIn("ROUND", values["慕课20分制"])
        self.assertIn("ROUND", values["讨论10分制"])

    def test_multiple_cl_files_merge_duplicates_and_keep_sources(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            cl_1 = tmp_path / "CL_001.xlsx"
            cl_2 = tmp_path / "CL_002.xlsx"
            pd.DataFrame(
                [
                    ["学号", "姓名", "班级"],
                    ["U202400001", "张三", "计2401"],
                    ["U202400002", "李四", "计2401"],
                ]
            ).to_excel(cl_1, header=False, index=False)
            pd.DataFrame(
                [
                    ["学号", "姓名", "班级"],
                    ["U202400001", "张三", "计2401"],
                    ["U202400003", "王五", "计2402"],
                ]
            ).to_excel(cl_2, header=False, index=False)

            students, _, report = extract_rosters([cl_1, cl_2])

        self.assertEqual(len(students), 3)
        self.assertEqual(len(report), 2)
        self.assertIn("CL_002.xlsx", students["U202400001"].duplicate_records[0])


if __name__ == "__main__":
    unittest.main()
